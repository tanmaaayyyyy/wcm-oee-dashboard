"""
WCM Loss Analysis & OEE Monitoring Dashboard
Production data generator.

Generates a realistic 90 day production line dataset for a dairy and beverage
carton packaging operation and writes it to data/production_data.xlsx with two
sheets:

  1. "Production Data" : raw daily shift level records with OEE components
  2. "KPI Summary"     : aggregated KPIs for reporting and Power BI validation

The data is engineered to tell a clear WCM (World Class Manufacturing) story:
  - Overall OEE averages roughly 72 to 78 percent (world class target is 85)
  - Line A is the best performing line
  - Night shift carries slightly more downtime
  - Mechanical Failure is the dominant loss category so the Pareto is clear

Run:
    python generate_data.py
"""

from __future__ import annotations

import os
from datetime import date, timedelta

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------- #
# Configuration constants
# --------------------------------------------------------------------------- #

RANDOM_SEED = 42
NUM_DAYS = 90
PLANNED_PRODUCTION_TIME_HRS = 8.0
WORLD_CLASS_OEE = 0.85

SHIFTS = ["Morning", "Afternoon", "Night"]
LINES = ["Line A", "Line B", "Line C"]

DOWNTIME_CATEGORIES = [
    "Mechanical Failure",
    "Changeover",
    "Material Shortage",
    "Planned Maintenance",
    "Quality Hold",
]

REJECTION_CATEGORIES = [
    "Sealing Defect",
    "Label Misalignment",
    "Weight Variance",
    "Contamination",
    "Dimension Error",
]

OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "production_data.xlsx")

# Per line tuning knobs. Line A is engineered to be the strongest performer:
# lower base downtime, higher line speed, better performance and quality.
LINE_PROFILES = {
    "Line A": {"downtime_base": 0.80, "speed": 1150, "perf_center": 0.93, "reject_rate": 0.018},
    "Line B": {"downtime_base": 1.10, "speed": 1000, "perf_center": 0.89, "reject_rate": 0.026},
    "Line C": {"downtime_base": 1.35, "speed": 900, "perf_center": 0.86, "reject_rate": 0.033},
}

# Night shift adds downtime; afternoon a touch; morning is the cleanest run.
SHIFT_DOWNTIME_ADJ = {"Morning": -0.10, "Afternoon": 0.10, "Night": 0.45}

# Weighted so Mechanical Failure dominates the Pareto, then Changeover, etc.
DOWNTIME_WEIGHTS = [0.42, 0.23, 0.16, 0.11, 0.08]
REJECTION_WEIGHTS = [0.38, 0.24, 0.17, 0.12, 0.09]


# --------------------------------------------------------------------------- #
# Data generation
# --------------------------------------------------------------------------- #


def generate_records(rng: np.random.Generator) -> pd.DataFrame:
    """Build the raw production records as a DataFrame."""
    start_date = date.today() - timedelta(days=NUM_DAYS - 1)
    rows: list[dict] = []

    for day_offset in range(NUM_DAYS):
        current_date = start_date + timedelta(days=day_offset)
        for line in LINES:
            profile = LINE_PROFILES[line]
            for shift in SHIFTS:
                rows.append(
                    _build_row(rng, current_date, line, shift, profile)
                )

    return pd.DataFrame(rows)


def _build_row(rng, current_date, line, shift, profile) -> dict:
    """Generate a single shift level record for one line."""
    # --- Downtime (hours) -------------------------------------------------- #
    downtime_mean = profile["downtime_base"] + SHIFT_DOWNTIME_ADJ[shift]
    downtime = rng.normal(loc=downtime_mean, scale=0.45)
    downtime = float(np.clip(downtime, 0.0, 2.5))
    downtime = round(downtime, 2)

    downtime_category = rng.choice(DOWNTIME_CATEGORIES, p=DOWNTIME_WEIGHTS)

    # --- Planned units (line speed driven) --------------------------------- #
    speed = profile["speed"]
    units_planned = int(rng.normal(loc=speed, scale=40))
    units_planned = int(np.clip(units_planned, 800, 1200))

    # --- Performance and produced units ------------------------------------ #
    # Performance reflects speed / minor stop losses only. Downtime is captured
    # separately by Availability, so it is deliberately NOT folded in here,
    # otherwise OEE = Availability x Performance x Quality would double count
    # the downtime loss (Performance = Units_Produced / Units_Planned).
    perf_ratio = rng.normal(loc=profile["perf_center"], scale=0.03)
    perf_ratio = float(np.clip(perf_ratio, 0.78, 0.99))
    units_produced = int(units_planned * perf_ratio)
    units_produced = max(units_produced, 0)

    # --- Rejections -------------------------------------------------------- #
    reject_rate = rng.normal(loc=profile["reject_rate"], scale=0.006)
    reject_rate = float(np.clip(reject_rate, 0.004, 0.06))
    units_rejected = int(round(units_produced * reject_rate))
    units_rejected = min(units_rejected, units_produced)

    rejection_category = rng.choice(REJECTION_CATEGORIES, p=REJECTION_WEIGHTS)

    return {
        "Date": current_date,
        "Shift": shift,
        "Line": line,
        "Planned_Production_Time_hrs": PLANNED_PRODUCTION_TIME_HRS,
        "Downtime_hrs": downtime,
        "Downtime_Category": downtime_category,
        "Units_Planned": units_planned,
        "Units_Produced": units_produced,
        "Units_Rejected": units_rejected,
        "Rejection_Category": rejection_category,
    }


def add_oee_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Compute Availability, Performance, Quality and OEE."""
    out = df.copy()

    out["Availability"] = (
        (out["Planned_Production_Time_hrs"] - out["Downtime_hrs"])
        / out["Planned_Production_Time_hrs"]
    )
    out["Performance"] = out["Units_Produced"] / out["Units_Planned"]
    # Guard against divide by zero if a shift produced nothing.
    out["Quality"] = np.where(
        out["Units_Produced"] > 0,
        (out["Units_Produced"] - out["Units_Rejected"]) / out["Units_Produced"],
        0.0,
    )
    out["OEE"] = out["Availability"] * out["Performance"] * out["Quality"]

    for col in ["Availability", "Performance", "Quality", "OEE"]:
        out[col] = out[col].round(4)

    return out


# --------------------------------------------------------------------------- #
# KPI summary
# --------------------------------------------------------------------------- #


def build_kpi_summary(df: pd.DataFrame) -> dict[str, pd.DataFrame]:
    """Build the aggregated KPI tables used in the KPI Summary sheet."""
    overall_oee = df["OEE"].mean()

    oee_by_line = (
        df.groupby("Line")["OEE"].mean().reset_index().sort_values("OEE", ascending=False)
    )
    oee_by_shift = (
        df.groupby("Shift")["OEE"].mean().reset_index().sort_values("OEE", ascending=False)
    )

    # Top 3 downtime categories by total hours.
    downtime = (
        df.groupby("Downtime_Category")["Downtime_hrs"].sum().reset_index()
    )
    total_downtime = downtime["Downtime_hrs"].sum()
    downtime["Percent"] = downtime["Downtime_hrs"] / total_downtime
    top_downtime = downtime.sort_values("Downtime_hrs", ascending=False).head(3)

    # Top 3 rejection categories by total count.
    rejects = (
        df.groupby("Rejection_Category")["Units_Rejected"].sum().reset_index()
    )
    total_rejects = rejects["Units_Rejected"].sum()
    rejects["Percent"] = rejects["Units_Rejected"] / total_rejects
    top_rejects = rejects.sort_values("Units_Rejected", ascending=False).head(3)

    # Month over month OEE trend (last 3 calendar months present in data).
    trend = df.copy()
    trend["Month"] = pd.to_datetime(trend["Date"]).dt.to_period("M").astype(str)
    monthly = trend.groupby("Month")["OEE"].mean().reset_index().tail(3)

    return {
        "overall_oee": overall_oee,
        "oee_by_line": oee_by_line,
        "oee_by_shift": oee_by_shift,
        "top_downtime": top_downtime,
        "top_rejects": top_rejects,
        "monthly": monthly,
    }


# --------------------------------------------------------------------------- #
# Excel writing and formatting
# --------------------------------------------------------------------------- #

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SECTION_FILL = PatternFill("solid", fgColor="2E75B6")
TOTAL_FILL = PatternFill("solid", fgColor="DDEBF7")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FONT = Font(bold=True, color="FFFFFF", size=12)
TITLE_FONT = Font(bold=True, size=16, color="1F4E78")
BOLD_FONT = Font(bold=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


def write_excel(df: pd.DataFrame, kpis: dict) -> None:
    """Write both sheets and apply professional formatting."""
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Production Data", index=False)
        # KPI sheet is built cell by cell for full formatting control.
        writer.book.create_sheet("KPI Summary")
        _format_production_sheet(writer.book["Production Data"], df)
        _write_kpi_sheet(writer.book["KPI Summary"], kpis)


def _format_production_sheet(ws, df: pd.DataFrame) -> None:
    """Style the raw data sheet: header band, widths, number formats, freeze."""
    n_cols = len(df.columns)

    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER

    # Column widths sized to header text.
    for col_idx, col_name in enumerate(df.columns, start=1):
        width = max(len(str(col_name)) + 2, 12)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Percentage formatting for the four ratio columns.
    pct_cols = {"Availability", "Performance", "Quality", "OEE"}
    col_index = {name: i + 1 for i, name in enumerate(df.columns)}
    last_row = ws.max_row

    for name in pct_cols:
        letter = get_column_letter(col_index[name])
        for row in range(2, last_row + 1):
            ws[f"{letter}{row}"].number_format = "0.0%"

    # Date column formatting.
    date_letter = get_column_letter(col_index["Date"])
    for row in range(2, last_row + 1):
        ws[f"{date_letter}{row}"].number_format = "yyyy-mm-dd"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _write_kpi_sheet(ws, kpis: dict) -> None:
    """Lay out the KPI Summary sheet section by section."""
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 14

    # Title.
    ws["A1"] = "WCM Loss Analysis & OEE  |  KPI Summary"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")
    ws["A1"].alignment = LEFT
    ws.row_dimensions[1].height = 24

    ws["A2"] = f"World Class OEE Target: {WORLD_CLASS_OEE:.0%}"
    ws["A2"].font = Font(italic=True, color="808080")
    ws.merge_cells("A2:D2")

    row = 4
    row = _kpi_headline(ws, row, "Overall OEE Average", kpis["overall_oee"])
    row += 1

    row = _kpi_table(
        ws, row, "OEE by Line",
        kpis["oee_by_line"], ["Line", "OEE"], pct_cols={"OEE"},
    )
    row += 1

    row = _kpi_table(
        ws, row, "OEE by Shift",
        kpis["oee_by_shift"], ["Shift", "OEE"], pct_cols={"OEE"},
    )
    row += 1

    dt = kpis["top_downtime"].rename(
        columns={"Downtime_Category": "Category", "Downtime_hrs": "Hours", "Percent": "% of Total"}
    )
    dt["Hours"] = dt["Hours"].round(1)
    row = _kpi_table(
        ws, row, "Top 3 Downtime Categories",
        dt, ["Category", "Hours", "% of Total"], pct_cols={"% of Total"},
    )
    row += 1

    rj = kpis["top_rejects"].rename(
        columns={"Rejection_Category": "Category", "Units_Rejected": "Count", "Percent": "% of Total"}
    )
    row = _kpi_table(
        ws, row, "Top 3 Rejection Categories",
        rj, ["Category", "Count", "% of Total"], pct_cols={"% of Total"},
    )
    row += 1

    mo = kpis["monthly"].rename(columns={"Month": "Month", "OEE": "OEE"})
    row = _kpi_table(
        ws, row, "Month over Month OEE Trend (Last 3 Months)",
        mo, ["Month", "OEE"], pct_cols={"OEE"},
    )


def _kpi_headline(ws, row: int, label: str, value: float) -> int:
    """Render a single large headline metric."""
    ws.cell(row=row, column=1, value=label)
    ws.cell(row=row, column=1).font = SECTION_FONT
    ws.cell(row=row, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)

    val_cell = ws.cell(row=row, column=3, value=value)
    val_cell.number_format = "0.0%"
    val_cell.font = Font(bold=True, size=14, color="1F4E78")
    val_cell.fill = TOTAL_FILL
    val_cell.alignment = CENTER
    val_cell.border = BORDER
    return row + 1


def _kpi_table(ws, row: int, title: str, frame: pd.DataFrame,
               columns: list[str], pct_cols: set[str]) -> int:
    """Render a titled KPI table starting at `row`. Returns next free row."""
    # Section title bar.
    ws.cell(row=row, column=1, value=title)
    ws.cell(row=row, column=1).font = SECTION_FONT
    ws.cell(row=row, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(columns))
    row += 1

    # Column headers.
    for c_idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=c_idx, value=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER
    row += 1

    # Data rows.
    for _, record in frame.iterrows():
        for c_idx, col in enumerate(columns, start=1):
            value = record[col]
            cell = ws.cell(row=row, column=c_idx, value=value)
            cell.border = BORDER
            if col in pct_cols:
                cell.number_format = "0.0%"
                cell.alignment = CENTER
            elif isinstance(value, (int, float)) and not isinstance(value, bool):
                cell.number_format = "#,##0.0" if isinstance(value, float) else "#,##0"
                cell.alignment = CENTER
            else:
                cell.alignment = LEFT
        row += 1

    return row


# --------------------------------------------------------------------------- #
# Entry point
# --------------------------------------------------------------------------- #


def main() -> None:
    rng = np.random.default_rng(RANDOM_SEED)

    raw = generate_records(rng)
    data = add_oee_columns(raw)
    kpis = build_kpi_summary(data)

    write_excel(data, kpis)

    print("Generated production dataset")
    print(f"  Rows            : {len(data)}")
    print(f"  Date range      : {data['Date'].min()} to {data['Date'].max()}")
    print(f"  Overall OEE     : {kpis['overall_oee']:.1%}")
    print("  OEE by Line     :")
    for _, r in kpis["oee_by_line"].iterrows():
        print(f"      {r['Line']:<8}: {r['OEE']:.1%}")
    print("  OEE by Shift    :")
    for _, r in kpis["oee_by_shift"].iterrows():
        print(f"      {r['Shift']:<10}: {r['OEE']:.1%}")
    print("  Top downtime    :")
    for _, r in kpis["top_downtime"].iterrows():
        print(f"      {r['Downtime_Category']:<20}: {r['Downtime_hrs']:.0f} hrs ({r['Percent']:.1%})")
    print(f"  Saved to        : {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
